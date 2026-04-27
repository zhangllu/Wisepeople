#!/usr/bin/env python3
"""
Extract 通识千书 data from Excel to structured JSON files.

Reads `notes/通识千书书单-聪明的阅读者.xlsx` (2 sheets) and outputs:
  - src/data/books.json        — 1197 books
  - src/data/authors.json      — 778 unique authors
  - src/data/topics.json       — 51 sub-topics (元典 + 50 themes)
  - src/data/questions.json    — 10 big questions
  - src/data/minimum_books.json — 56 books with summaries

Usage:
  cd project-root
  uv run python scripts/extract_books_data.py
"""

import hashlib
import json
import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = PROJECT_ROOT / "notes" / "通识千书书单-聪明的阅读者.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "src" / "data"

# ── Canonical data from 通识千书框架目录.md ─────────────────────────────

QUESTIONS_DATA = [
    {
        "code": "Q01",
        "number": 1,
        "title": "知识的知识",
        "dimension": "元问题",
        "subtitle": "所有大问题的前置问题——关于知识本身的知识。认识论、方法论与创新学。",
        "representative_sages": "培根、笛卡儿、休谟、康德、胡塞尔、维特根斯坦、波兰尼、波利亚、哈代、伽德纳、皮尔斯、艾柯、坎贝尔、布鲁纳、王阳明、贝特森、契克森米哈赖、诺曼、野中郁次郎",
    },
    {
        "code": "Q02",
        "number": 2,
        "title": "如何理解世界",
        "dimension": "天",
        "subtitle": "世界观——认识我们所处的物质世界与生命世界。",
        "representative_sages": "",
    },
    {
        "code": "Q03",
        "number": 3,
        "title": "如何理解历史",
        "dimension": "天",
        "subtitle": "从历史中汲取智慧——理解文明的兴衰脉络。",
        "representative_sages": "",
    },
    {
        "code": "Q04",
        "number": 4,
        "title": "如何理解时代",
        "dimension": "地",
        "subtitle": "时代诊断——识别我们所处时代的核心挑战与危机。",
        "representative_sages": "",
    },
    {
        "code": "Q05",
        "number": 5,
        "title": "如何理解社会",
        "dimension": "地",
        "subtitle": "人类社会的基本运作逻辑——交换、交易、博弈与多样性。",
        "representative_sages": "",
    },
    {
        "code": "Q06",
        "number": 6,
        "title": "如何理解组织",
        "dimension": "地",
        "subtitle": "现代社会的组织逻辑——如何创建和管理组织。",
        "representative_sages": "",
    },
    {
        "code": "Q07",
        "number": 7,
        "title": "如何理解家庭",
        "dimension": "地",
        "subtitle": "人生最基本的亲密关系单元——从建立到告别。",
        "representative_sages": "",
    },
    {
        "code": "Q08",
        "number": 8,
        "title": "如何理解人性",
        "dimension": "人",
        "subtitle": "人何以为人——从生理基础到文学表达的人性探索。",
        "representative_sages": "",
    },
    {
        "code": "Q09",
        "number": 9,
        "title": "如何理解身体",
        "dimension": "人",
        "subtitle": "身体是灵魂的居所——身心健康与环境适应。",
        "representative_sages": "",
    },
    {
        "code": "Q10",
        "number": 10,
        "title": "如何理解信仰",
        "dimension": "人",
        "subtitle": "超越性追问——理解不同文明的精神传统。",
        "representative_sages": "",
    },
]

# Canonical sub-topic definitions from the framework directory
SUB_TOPIC_DEFS = {
    "0": {
        "title": "元典：人类文明十三经",
        "core_field": "人类文明源头",
        "representative_discipline": "经典",
        "question_number": 0,
    },
    "1.1": {"title": "总论", "core_field": "认识论、知识论", "representative_discipline": "哲学", "question_number": 1},
    "1.2": {"title": "笛卡儿信徒", "core_field": "数学思维、问题解决", "representative_discipline": "数学", "question_number": 1},
    "1.3": {"title": "讲故事的人", "core_field": "符号系统、叙事结构", "representative_discipline": "符号学、叙事学、修辞学", "question_number": 1},
    "1.4": {"title": "达尔文信徒", "core_field": "行动哲学、实践智慧", "representative_discipline": "行动科学", "question_number": 1},
    "1.5": {"title": "成为创新者", "core_field": "创造力、设计思维", "representative_discipline": "创新学、设计学", "question_number": 1},
    "2.1": {"title": "总论", "core_field": "世界观、本体论", "representative_discipline": "哲学", "question_number": 2},
    "2.2": {"title": "进化论与复杂科学", "core_field": "演化、复杂系统", "representative_discipline": "进化生物学、复杂科学", "question_number": 2},
    "2.3": {"title": "宇宙天文、山川地理", "core_field": "宇宙、地球、生态", "representative_discipline": "天文学、地理学、生态学", "question_number": 2},
    "2.4": {"title": "物质规律", "core_field": "物理、化学基本规律", "representative_discipline": "物理学、化学", "question_number": 2},
    "2.5": {"title": "生物奥秘", "core_field": "生命科学", "representative_discipline": "生物学、神经科学", "question_number": 2},
    "3.1": {"title": "总论", "core_field": "史学理论、历史哲学", "representative_discipline": "历史学", "question_number": 3},
    "3.2": {"title": "中国通史", "core_field": "中国文明通史", "representative_discipline": "中国史", "question_number": 3},
    "3.3": {"title": "世界通史", "core_field": "全球文明通史", "representative_discipline": "世界史", "question_number": 3},
    "3.4": {"title": "中国断代史与专题史", "core_field": "中国各朝代与专题", "representative_discipline": "中国断代史", "question_number": 3},
    "3.5": {"title": "西方断代史与专题史", "core_field": "西方各时期与专题", "representative_discipline": "西方史", "question_number": 3},
    "4.1": {"title": "总论", "core_field": "时代诊断、文明批判", "representative_discipline": "社会学、哲学", "question_number": 4},
    "4.2": {"title": "直接暴力", "core_field": "军事、法律与控制", "representative_discipline": "军事学、法学", "question_number": 4},
    "4.3": {"title": "间接暴力", "core_field": "结构性暴力、制度不公", "representative_discipline": "政治学、社会学", "question_number": 4},
    "4.4": {"title": "互联时代", "core_field": "信息、网络与数字社会", "representative_discipline": "传播学、网络科学", "question_number": 4},
    "4.5": {"title": "对抗异化", "core_field": "人的异化与解放", "representative_discipline": "批判理论、文化研究", "question_number": 4},
    "5.1": {"title": "总论", "core_field": "社会学基础", "representative_discipline": "社会学", "question_number": 5},
    "5.2": {"title": "社会交换", "core_field": "社会网络、信任、合作", "representative_discipline": "社会心理学", "question_number": 5},
    "5.3": {"title": "经济交易", "core_field": "市场、价值、分配", "representative_discipline": "经济学", "question_number": 5},
    "5.4": {"title": "政治博弈", "core_field": "权力、制度、治理", "representative_discipline": "政治学", "question_number": 5},
    "5.5": {"title": "人类的多样性", "core_field": "文化、族群、性别", "representative_discipline": "人类学、文化研究", "question_number": 5},
    "6.1": {"title": "总论", "core_field": "组织理论", "representative_discipline": "组织行为学、管理学", "question_number": 6},
    "6.2": {"title": "组织中的人", "core_field": "人力资源、领导力", "representative_discipline": "组织心理学", "question_number": 6},
    "6.3": {"title": "组织中的钱", "core_field": "财务、会计、资本", "representative_discipline": "财务学", "question_number": 6},
    "6.4": {"title": "组织中的事", "core_field": "运营、项目、流程", "representative_discipline": "运营管理", "question_number": 6},
    "6.5": {"title": "创建组织", "core_field": "创业、企业家精神", "representative_discipline": "创业学", "question_number": 6},
    "7.1": {"title": "总论", "core_field": "家庭社会学", "representative_discipline": "社会学、心理学", "question_number": 7},
    "7.2": {"title": "创建家庭：婚姻与恋爱", "core_field": "亲密关系", "representative_discipline": "社会心理学", "question_number": 7},
    "7.3": {"title": "扩大家庭：生儿育女", "core_field": "育儿、教育", "representative_discipline": "发展心理学、教育学", "question_number": 7},
    "7.4": {"title": "改善家庭：家庭治疗", "core_field": "家庭沟通、修复", "representative_discipline": "家庭治疗", "question_number": 7},
    "7.5": {"title": "离开家庭：安度晚年与告别世界", "core_field": "老龄化、死亡", "representative_discipline": "老年学、临终关怀", "question_number": 7},
    "8.1": {"title": "总论", "core_field": "人性论", "representative_discipline": "哲学、心理学", "question_number": 8},
    "8.2": {"title": "人性的生理基础", "core_field": "大脑、基因、进化", "representative_discipline": "神经科学、进化心理学", "question_number": 8},
    "8.3": {"title": "人性的心理学理解", "core_field": "认知、情感、行为", "representative_discipline": "心理学", "question_number": 8},
    "8.4": {"title": "人性的语言学理解", "core_field": "语言、思维、文化", "representative_discipline": "语言学", "question_number": 8},
    "8.5": {"title": "人性的文学理解", "core_field": "文学中的人性洞察", "representative_discipline": "文学", "question_number": 8},
    "9.1": {"title": "总论", "core_field": "身体哲学、医学哲学", "representative_discipline": "哲学、医学", "question_number": 9},
    "9.2": {"title": "生理健康", "core_field": "运动、营养、生活方式", "representative_discipline": "医学、运动科学", "question_number": 9},
    "9.3": {"title": "心理健康", "core_field": "情绪管理、心理调适", "representative_discipline": "临床心理学", "question_number": 9},
    "9.4": {"title": "环境健康", "core_field": "人与环境的关系", "representative_discipline": "环境科学", "question_number": 9},
    "9.5": {"title": "疾病与治疗", "core_field": "疾病认知、医疗决策", "representative_discipline": "医学", "question_number": 9},
    "10.1": {"title": "总论", "core_field": "宗教哲学、信仰研究", "representative_discipline": "宗教学、哲学", "question_number": 10},
    "10.2": {"title": "儒家", "core_field": "儒家传统", "representative_discipline": "中国哲学", "question_number": 10},
    "10.3": {"title": "东方宗教", "core_field": "佛教、道教、印度教", "representative_discipline": "东方哲学", "question_number": 10},
    "10.4": {"title": "西方宗教", "core_field": "基督教、伊斯兰教", "representative_discipline": "西方哲学", "question_number": 10},
    "10.5": {"title": "新道学", "core_field": "现代灵性探索", "representative_discipline": "比较宗教学", "question_number": 10},
}

# ── Helpers ─────────────────────────────────────────────────────────────


def extract_topic_code(raw_value: str) -> str:
    """Extract topic code from Excel '大问题' value.

    Examples:
      "0.元典十三经"        → "0"
      "1.1知识的知识"      → "1.1"
      "1.2 笛卡儿信徒"     → "1.2"
      "10.5 新道学"        → "10.5"
    """
    raw = str(raw_value).strip()
    m = re.match(r"(\d+(?:\.\d+)?)", raw)
    return m.group(1) if m else ""


def extract_topic_title(raw_value: str) -> str:
    """Extract the display title from Excel '大问题' value (strip code prefix)."""
    raw = str(raw_value).strip()
    m = re.match(r"\d+(?:\.\d+)?\s*", raw)
    if m:
        return raw[m.end():].strip()
    return raw


def make_slug(text: str) -> str:
    """Create an ASCII slug from Chinese/English text.

    For Chinese text, use a hash-based approach since pinyin is unavailable.
    For English text, lowercase and hyphenate.
    """
    slug = text.strip().lower()
    # If contains Chinese, append a short hash for uniqueness
    if any("\u4e00" <= c <= "\u9fff" for c in slug):
        h = hashlib.md5(text.encode("utf-8")).hexdigest()[:8]
        return f"b-{h}"
    # For English text: lowercase, keep letters/digits/hyphens
    slug = re.sub(r"[^a-z0-9]+", "-", slug).strip("-")
    return slug if slug else f"b-{hashlib.md5(text.encode('utf-8')).hexdigest()[:8]}"


def make_author_slug(name: str) -> str:
    """Create a slug for an author name."""
    name = name.strip()
    if any("\u4e00" <= c <= "\u9fff" for c in name):
        # Use a hash of the name for Chinese authors
        h = hashlib.md5(name.encode("utf-8")).hexdigest()[:8]
        return f"a-{h}"
    slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return slug if slug else f"a-{hashlib.md5(name.encode('utf-8')).hexdigest()[:8]}"


def clean_author_name(raw: str) -> str:
    """Clean an author name string.

    Strips whitespace/punctuation, removes extra spaces, keeps separators.
    """
    return str(raw).strip() if raw else ""


# ── Main extraction ──────────────────────────────────────────────────────


def extract_sheet1(ws):
    """Extract books and authors from Sheet 1 (通识千书1000+全)."""
    books = []
    author_map = {}  # normalized_name -> {slug, name, book_slugs, topic_codes_set}

    max_row = ws.max_row

    for r in range(2, max_row + 1):
        raw_question = ws.cell(r, 4).value  # column D: 大问题
        title = ws.cell(r, 5).value  # column E: 书名
        author_raw = ws.cell(r, 6).value  # column F: 作者
        year = ws.cell(r, 7).value  # column G: 出版年
        publisher = ws.cell(r, 8).value  # column H: 出版社
        link = ws.cell(r, 9).value  # column I: 链接-link
        tag = ws.cell(r, 12).value  # column L: 标签

        if not title:
            continue

        title = str(title).strip()
        if not title:
            continue

        author_name = clean_author_name(author_raw) if author_raw else "佚名"

        topic_code = extract_topic_code(raw_question) if raw_question else ""

        # Generate book slug and ID
        book_slug = make_slug(f"{title}-{author_name}")
        book_id = r - 1  # sequential ID starting from 1

        book = {
            "id": book_id,
            "slug": book_slug,
            "title": title,
            "author": author_name,
            "authorSlug": make_author_slug(author_name),
            "year": int(year) if year and str(year).isdigit() else None,
            "publisher": str(publisher).strip() if publisher else "",
            "doubanLink": str(link).strip() if link else "",
            "topicCode": topic_code,
            "tags": str(tag).strip() if tag else "",
        }
        books.append(book)

        # Track author
        if author_name not in author_map:
            author_map[author_name] = {
                "slug": make_author_slug(author_name),
                "name": author_name,
                "bookSlugs": [],
                "topicCodes": set(),
            }
        author_map[author_name]["bookSlugs"].append(book_slug)
        if topic_code:
            author_map[author_name]["topicCodes"].add(topic_code)

    # Convert authors to list
    authors = []
    for name, data in author_map.items():
        authors.append({
            "slug": data["slug"],
            "name": data["name"],
            "bookSlugs": data["bookSlugs"],
            "topicCodes": sorted(data["topicCodes"]),
        })

    # Sort authors by name for determinism
    authors.sort(key=lambda a: a["name"])

    return books, authors


def extract_sheet2(ws):
    """Extract minimum book list from Sheet 2 (最小书单-含摘要)."""
    min_books = []
    max_row = ws.max_row

    for r in range(2, max_row + 1):
        isbn = ws.cell(r, 1).value  # column A: ISBN
        title = ws.cell(r, 2).value  # column B: 标题
        author = ws.cell(r, 3).value  # column C: 作者
        link = ws.cell(r, 5).value  # column E: 链接-Url
        publisher = ws.cell(r, 8).value  # column H: 出版社
        tag_class = ws.cell(r, 9).value  # column I: 标签分类
        translator = ws.cell(r, 11).value  # column K: 翻译者
        summary = ws.cell(r, 12).value  # column L: 摘要笔记

        if not title:
            continue

        title = str(title).strip()
        if not title:
            continue

        author_name = clean_author_name(author) if author else ""

        # Extract <a> tags etc from summary, clean it
        summary_text = ""
        if summary:
            summary_text = str(summary).strip()
            # Remove HTML tags
            summary_text = re.sub(r"<[^>]+>", "", summary_text)
            # Clean up multiple whitespace
            summary_text = re.sub(r"\s+", "", summary_text)

        min_books.append({
            "isbn": str(isbn).strip() if isbn else "",
            "title": title,
            "author": author_name,
            "slug": make_slug(f"{title}-{author_name}"),
            "translator": str(translator).strip() if translator else "",
            "publisher": str(publisher).strip() if publisher else "",
            "doubanLink": str(link).strip() if link else "",
            "tagClass": str(tag_class).strip() if tag_class else "",
            "summary": summary_text,
        })

    return min_books


def build_topics(excel_topic_codes_seen):
    """Build topics list, merging canonical definitions with Excel-derived titles.

    For sub-topics defined in the framework, use the canonical title/field/discipline.
    For any code not in the framework (unlikely), derive from Excel.
    """
    topics = []
    question_number_to_topic_codes = defaultdict(list)

    # Determine which codes exist
    all_codes = sorted(excel_topic_codes_seen, key=lambda c: [int(x) if x else 0 for x in c.split(".")])

    for code in all_codes:
        if code in SUB_TOPIC_DEFS:
            defn = SUB_TOPIC_DEFS[code]
        else:
            # Fallback: code not in framework (shouldn't happen)
            defn = {"title": code, "core_field": "", "representative_discipline": ""}

        q_num = defn.get("question_number", 0)
        question_number_to_topic_codes[q_num].append(code)

        topics.append({
            "code": code,
            "title": defn["title"],
            "coreField": defn["core_field"],
            "representativeDiscipline": defn["representative_discipline"],
            "questionNumber": q_num,
        })

    return topics, dict(question_number_to_topic_codes)


def build_questions(qnum_to_topic_codes):
    """Build questions list, attaching sub-topic codes."""
    questions = []
    for qd in QUESTIONS_DATA:
        qnum = qd["number"]
        sub_topic_codes = sorted(qnum_to_topic_codes.get(qnum, []))
        questions.append({
            "code": qd["code"],
            "number": qd["number"],
            "title": qd["title"],
            "dimension": qd["dimension"],
            "subtitle": qd["subtitle"],
            "subTopicCodes": sub_topic_codes,
        })
    return questions


def write_json(data, filename):
    """Write data to JSON file with indentation."""
    path = OUTPUT_DIR / filename
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  ✓ Wrote {len(data)} records → {path}")


def main():
    print("=" * 60)
    print("通识千书 数据提取脚本")
    print("=" * 60)

    # Ensure output directory exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Open workbook
    print(f"\n📖 Reading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # ── Sheet 1: 通识千书1000+全 ──
    ws1 = wb["通识千书1000+全"]
    print(f"  Sheet 1: {ws1.max_row} rows × {ws1.max_column} cols")

    books, authors = extract_sheet1(ws1)
    print(f"  → {len(books)} books, {len(authors)} unique authors")

    # Collect all topic codes seen in Excel
    excel_topic_codes = set()
    for b in books:
        if b["topicCode"]:
            excel_topic_codes.add(b["topicCode"])

    # ── Sheet 2: 最小书单-含摘要 ──
    ws2 = wb["最小书单-含摘要"]
    print(f"\n  Sheet 2: {ws2.max_row} rows × {ws2.max_column} cols")

    minimum_books = extract_sheet2(ws2)
    print(f"  → {len(minimum_books)} minimum books with summaries")

    # Build topics and questions
    print(f"\n  → {len(excel_topic_codes)} unique topic codes found in Excel")
    topics, qnum_to_topic_codes = build_topics(excel_topic_codes)
    questions = build_questions(qnum_to_topic_codes)

    # ── Write output files ──
    print(f"\n📝 Writing JSON files to {OUTPUT_DIR}/")
    write_json(books, "books.json")
    write_json(authors, "authors.json")
    write_json(topics, "topics.json")
    write_json(questions, "questions.json")
    write_json(minimum_books, "minimum_books.json")

    # ── Summary ──
    print(f"\n{'=' * 60}")
    print("📊 提取完成")
    print(f"{'=' * 60}")
    print(f"  书籍 (Books):              {len(books)}")
    print(f"  作者 (Authors):            {len(authors)}")
    print(f"  子主题 (Topics):           {len(topics)}")
    print(f"  大问题 (Questions):        {len(questions)}")
    print(f"  最小书单 (Min Books):      {len(minimum_books)}")

    # Validate topic coverage
    topic_codes = {t["code"] for t in topics}
    expected_codes = set(SUB_TOPIC_DEFS.keys())
    missing = expected_codes - topic_codes
    extra = topic_codes - expected_codes
    if missing:
        print(f"  ⚠ 框架有但Excel中缺失的子主题: {missing}")
    if extra:
        print(f"  ⚠ Excel中有但框架未定义的子主题: {extra}")
    if not missing and not extra:
        print(f"  ✅ 子主题覆盖完整: {len(expected_codes)}/{len(expected_codes)}")

    # Print a few samples
    print(f"\n📌 示例数据:")
    print(f"  Top 3 books:     {[b['title'] for b in books[:3]]}")
    print(f"  Top 3 authors:   {[a['name'] for a in authors[:3]]}")
    print(f"  Topic codes:     {sorted(t['code'] for t in topics)}")
    q_titles = [q["title"] for q in questions]
    print(f"  10 questions:    {q_titles}")
    print(f"  Min book titles: {[m['title'] for m in minimum_books[:3]]}")

    print(f"\n✅ 完成！\n")


if __name__ == "__main__":
    main()
